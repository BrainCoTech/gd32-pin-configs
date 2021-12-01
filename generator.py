from openpyxl import load_workbook

# Remove white space from workbook
def remove_whilespace(s):
	for i in range(1, s.max_row + 1):
		for j in range(1, s.max_column + 1):
			c = s.cell(row = i, column = j)
			if c.value:
				c.value = ''.join(str(c.value).split())

# Load workbook wrapper
def load_af_summary(fn):
	wb = load_workbook(fn)
	remove_whilespace(wb.worksheets[0])
	return wb

'''
Format signal-configs and pins alternate function definitions.

Return format:
{
	signal-configs: { (optional)

	},
	pins: {
		'PA0': {
			'pincodes': ['A', 'B', 'C'],
			'afs': {
				'modex': 0,
				'modey': 1
			}
		},
		'PA0': {
			'pincodes': ['B', 'C'],
			'afs': {
				'modey': 0,
				'modez': 1
			}
		}
	}
}
'''
def pinctrl_config_construct(af_summary, package_pins, exclude_memories_groups = None):
	pinctrl_config = {}
	signal_configs = {}
	pins_cfg = {}
	s = af_summary.worksheets[0]
	for i in range(2, s.max_row + 1):
		pin_name = s.cell(row = i, column = 1)

		afs = {}
		for j in range(2, s.max_column + 1):
			c = s.cell(row = i, column = j)
			# check cell is empty?
			if not c.value:
				continue

			modes = str(c.value).replace(',', '/')
			modes = modes.replace(')', ')/')
			for mode in modes.split('/'):
				if not mode:
					continue

				# exclude memories defined for this mode?
				if exclude_memories_groups and mode.find(')') > 0:
					mode_cfg = mode.replace('(', ')').split(')')
					mode = mode_cfg[0]
					signal_configs[mode] = {'exclude-memories': exclude_memories_groups[mode_cfg[1]]}

				afs[mode] = j - 2
	
		# No af defined for this pin
		if not afs:
			continue

		pincodes = []
		for package in package_pins:
			if pin_name.value in package_pins[package]:
				pincodes.append(package)

		pins_cfg[pin_name.value] = {'pincodes': pincodes,'afs': afs}

	if signal_configs:
			pinctrl_config['signal-configs'] = signal_configs
	pinctrl_config['pins'] = pins_cfg

	return pinctrl_config

def _signal_config_format(signal_configs):
	ws = '  '
	signals_yaml = 'signal-configs:\n'
	for signal in signal_configs:
		signal_yaml = ws * 1 + signal + ':\n'
		signal_config = signal_configs[signal]
		if 'exclude-memories' in signal_config:
			exclude_memory = ws * 2 + 'exclude-memories: ['
			for memory in signal_config['exclude-memories']:
				exclude_memory += str(memory) + ', '
			exclude_memory = exclude_memory[:-2] + ']\n'
			signal_yaml += exclude_memory
		signals_yaml += signal_yaml	

	return signals_yaml

def _pins_config_format(pins_config):
	ws = '  '
	pins_yaml = 'pins:\n'
	for pin in pins_config:
		pin_yaml = ws * 1 + pin + ':\n'

		pin_config = pins_config[pin]
		pincodes = ws * 2 + 'pincodes: ['
		for package in pin_config['pincodes']:
			pincodes += package + ', '
		pincodes = pincodes[:-2] + ']\n'
		pin_yaml += pincodes

		afs = ws * 2  + 'afs:\n'
		for mode in pin_config['afs']:
			afs += ws * 3 + mode + ': ' + str(pin_config['afs'][mode]) + '\n'
		pin_yaml += afs

		pins_yaml += pin_yaml

	return pins_yaml

# Generate pinctrl config to yaml
def pinctrl_config_format(pinctrl_config):
	pinctrl_yaml = ''
	if 'signal-configs' in pinctrl_config:
		pinctrl_yaml += _signal_config_format(pinctrl_config['signal-configs'])

	pinctrl_yaml += '\n'

	pinctrl_yaml += _pins_config_format(pinctrl_config['pins'])

	return pinctrl_yaml
