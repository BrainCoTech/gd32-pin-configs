from openpyxl import load_workbook
import re

def parse_pin_af_summary(filename):
    '''
    Parse pin alternate functions summary from file
    Input: pin definitions excel table copy from datasheet.
    Output: dict paired with pin and alternate. 
    '''
    wb = load_workbook(filename)
    sheet = wb.active
    summary = {}

    for i in range(2, sheet.max_row+1):
        pin = sheet.cell(row=i, column=1).value
        # The 'EVENTOUT' func is available for each pins.
        summary[pin] = []

        for j in range(2, sheet.max_column+1):
            func = sheet.cell(row=i, column=j).value
            if not func:
                summary[pin].append('')
                continue

            # Remove the note index.
            func = re.sub('\([0-9]+\)', '', func) 
            
            summary[pin].append(func)
    
    return summary

def parse_pin_definitions(filename):
    '''
     Parse pin definitions from excel table.
     Input: pin definitions excel table copy from datasheet.
     Return: dict paired with pin and alternate functions. 
    '''
    wb = load_workbook(filename)
    sheet = wb.active
    pin_function_definitions = {}
    default = ''
    alternate = ''
    additional = ''

    '''
        The pin function definitions located in column 5. start at row 2.
        Example value: 'Default:PC14Alternate:EVENTOUTAdditional:OSC32IN'
    '''
    for i in range(2, sheet.max_row+1):
        function_definition = sheet.cell(row=i, column=5).value
        function_definition = function_definition[len("Default:"):]

        tmp = function_definition.split("Alternate:")

        if len(tmp) == 2:
            default = tmp[0]
            function_definition = tmp[1]

            tmp = function_definition.split("Additional:")
            alternate = tmp[0]
            if len(tmp) == 2:
                additional = tmp[1]
            
            alternate = re.sub('\([0-9]+\)', '', alternate)
            alternate = re.sub('/', ',', alternate)

        else:
            tmp = function_definition.split("Additional:")
            default = tmp[0]
            if len(tmp) == 2:       
                additional = tmp[1]
    
        if re.match('P[A-F][0-9]+', default):
            pin = default
        elif additional and re.match('P[A-F][0-9]+', additional):
            # If pin in additional, it must be at the begin of additional string.
            pin = re.match('P[A-F][0-9]+', additional).group()
        else:
            ''' Not an I/O pin, ignored '''
            continue

        pin_function_definitions[pin] = alternate
    
    return pin_function_definitions
